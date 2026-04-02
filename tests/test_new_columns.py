"""
Tests for the new report columns added in the last two increments:

Component columns
─────────────────
 • Тип пакета / тип компонента   (package_type    from PURL)
 • PURL / технический идентификатор компонента (purl)
 • Признак принадлежности к поверхности атаки  (attack_surface)
 • Признак выполнения функций безопасности      (security_function)
 • Принадлежность к контейнерному образу        (container_image)
 • Роль компонента в составе контейнерного образа (container_role)

Vulnerability columns
─────────────────────
 • Рекомендация / компенсирующая мера           (recommendation)
 • Статус допустимости в рассматриваемой конфигурации (acceptability_status)
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict
from unittest.mock import patch

import pytest

from sbom_pipeline.vuln_merger import VulnFinding
from sbom_pipeline.exporter import Exporter, _COMP_COLUMNS, _VULN_COLUMNS

# ---------------------------------------------------------------------------
# Helpers / shared fixtures
# ---------------------------------------------------------------------------

def _sbom(**overrides) -> Dict[str, Any]:
    """Minimal valid CycloneDX SBOM dict."""
    base: Dict[str, Any] = {
        "bomFormat": "CycloneDX",
        "specVersion": "1.5",
        "components": [],
    }
    base.update(overrides)
    return base


def _comp(
    name: str = "libfoo",
    version: str = "1.0.0",
    purl: str = "pkg:pypi/libfoo@1.0.0",
    properties: list | None = None,
) -> Dict[str, Any]:
    c: Dict[str, Any] = {"type": "library", "name": name, "version": version, "purl": purl}
    if properties is not None:
        c["properties"] = properties
    return c


def _vuln(
    cve_id: str = "CVE-2024-0001",
    scanner: str = "trivy",
    recommendation: str = "",
    acceptability_status: str = "",
) -> VulnFinding:
    return VulnFinding(
        cve_id=cve_id,
        component_name="libfoo",
        component_version="1.0.0",
        component_purl="pkg:pypi/libfoo@1.0.0",
        cvss_score=7.5,
        severity="HIGH",
        description="Test vulnerability",
        scanner=scanner,
        fixed_version="1.1.0",
        recommendation=recommendation,
        acceptability_status=acceptability_status,
    )


# ===========================================================================
# 1. _purl_type helper
# ===========================================================================

class TestPurlType:
    """pipeline._purl_type() extracts the ecosystem from a PURL string."""

    def _purl_type(self, purl: str) -> str:
        from sbom_pipeline.pipeline import _purl_type
        return _purl_type(purl)

    def test_pypi(self):
        assert self._purl_type("pkg:pypi/requests@2.31.0") == "pypi"

    def test_maven(self):
        assert self._purl_type("pkg:maven/org.springframework/spring-core@5.3.0") == "maven"

    def test_npm(self):
        assert self._purl_type("pkg:npm/%40angular/core@15.0.0") == "npm"

    def test_apk(self):
        assert self._purl_type("pkg:apk/alpine/openssl@1.1.1k-r0") == "apk"

    def test_deb(self):
        assert self._purl_type("pkg:deb/debian/curl@7.64.0-4") == "deb"

    def test_empty(self):
        assert self._purl_type("") == ""

    def test_non_purl(self):
        assert self._purl_type("not-a-purl") == ""


# ===========================================================================
# 2. _find_prop helper
# ===========================================================================

class TestFindProp:
    """pipeline._find_prop() returns the first matching key value."""

    def _find_prop(self, props, keys):
        from sbom_pipeline.pipeline import _find_prop
        return _find_prop(props, keys)

    def test_first_key_wins(self):
        props = {"attack-surface": "yes", "attackSurface": "no"}
        assert self._find_prop(props, ("attack-surface", "attackSurface")) == "yes"

    def test_fallback_key(self):
        props = {"attackSurface": "yes"}
        assert self._find_prop(props, ("attack-surface", "attackSurface")) == "yes"

    def test_no_match_returns_empty(self):
        props = {"other": "value"}
        assert self._find_prop(props, ("attack-surface", "attackSurface")) == ""

    def test_empty_props(self):
        assert self._find_prop({}, ("attack-surface",)) == ""


# ===========================================================================
# 3. _extract_dependencies — component columns from SBOM
# ===========================================================================

class TestExtractDependencies:
    """pipeline._extract_dependencies() populates all new Dependency attributes."""

    def _extract(self, sbom: Dict[str, Any]) -> list:
        from sbom_pipeline.pipeline import _extract_dependencies
        # Patch Dependency._process_purl to skip HTTP calls
        with patch("sbom_pipeline.dependency.Dependency._process_purl"):
            return _extract_dependencies(sbom, "/fake/sbom.json")

    # -----------------------------------------------------------------------
    # package_type
    # -----------------------------------------------------------------------

    def test_package_type_pypi(self):
        sbom = _sbom(components=[_comp(purl="pkg:pypi/flask@3.0.0")])
        deps = self._extract(sbom)
        assert deps[0].package_type == "pypi"

    def test_package_type_maven(self):
        sbom = _sbom(components=[_comp(purl="pkg:maven/com.google.guava/guava@32.0.0")])
        deps = self._extract(sbom)
        assert deps[0].package_type == "maven"

    def test_package_type_apk(self):
        sbom = _sbom(components=[_comp(purl="pkg:apk/alpine/openssl@1.1.1k-r0")])
        deps = self._extract(sbom)
        assert deps[0].package_type == "apk"

    def test_package_type_empty_when_no_purl(self):
        comp = {"type": "library", "name": "unknown", "version": "0.0.1"}
        sbom = _sbom(components=[comp])
        deps = self._extract(sbom)
        assert deps[0].package_type == ""

    # -----------------------------------------------------------------------
    # purl
    # -----------------------------------------------------------------------

    def test_purl_is_preserved(self):
        purl = "pkg:npm/lodash@4.17.21"
        sbom = _sbom(components=[_comp(purl=purl)])
        deps = self._extract(sbom)
        assert deps[0].purl == purl

    # -----------------------------------------------------------------------
    # attack_surface
    # -----------------------------------------------------------------------

    @pytest.mark.parametrize("prop_name", [
        "attack-surface",
        "attack_surface",
        "attackSurface",
        "isAttackSurface",
    ])
    def test_attack_surface_recognised(self, prop_name: str):
        props = [{"name": prop_name, "value": "yes"}]
        sbom = _sbom(components=[_comp(properties=props)])
        deps = self._extract(sbom)
        assert deps[0].attack_surface == "yes"

    def test_attack_surface_empty_when_absent(self):
        sbom = _sbom(components=[_comp(properties=[])])
        deps = self._extract(sbom)
        assert deps[0].attack_surface == ""

    # -----------------------------------------------------------------------
    # security_function
    # -----------------------------------------------------------------------

    @pytest.mark.parametrize("prop_name", [
        "security-function",
        "security_function",
        "securityFunction",
        "isSecurityFunction",
    ])
    def test_security_function_recognised(self, prop_name: str):
        props = [{"name": prop_name, "value": "crypto"}]
        sbom = _sbom(components=[_comp(properties=props)])
        deps = self._extract(sbom)
        assert deps[0].security_function == "crypto"

    def test_security_function_empty_when_absent(self):
        sbom = _sbom(components=[_comp()])
        deps = self._extract(sbom)
        assert deps[0].security_function == ""

    # -----------------------------------------------------------------------
    # container_image
    # -----------------------------------------------------------------------

    def test_container_image_from_metadata(self):
        sbom = _sbom(
            metadata={"component": {"type": "container", "name": "my-image:latest"}},
            components=[_comp()],
        )
        deps = self._extract(sbom)
        assert deps[0].container_image == "my-image:latest"

    def test_container_image_empty_for_non_container_metadata(self):
        sbom = _sbom(
            metadata={"component": {"type": "application", "name": "my-app"}},
            components=[_comp()],
        )
        deps = self._extract(sbom)
        assert deps[0].container_image == ""

    def test_container_image_empty_when_no_metadata(self):
        sbom = _sbom(components=[_comp()])
        deps = self._extract(sbom)
        assert deps[0].container_image == ""

    # -----------------------------------------------------------------------
    # container_role
    # -----------------------------------------------------------------------

    @pytest.mark.parametrize("prop_name", [
        "container-role",
        "container_role",
        "containerRole",
        "cdx:docker:layer",
        "layer",
    ])
    def test_container_role_recognised(self, prop_name: str):
        props = [{"name": prop_name, "value": "os-packages"}]
        sbom = _sbom(components=[_comp(properties=props)])
        deps = self._extract(sbom)
        assert deps[0].container_role == "os-packages"

    def test_container_role_empty_when_absent(self):
        sbom = _sbom(components=[_comp()])
        deps = self._extract(sbom)
        assert deps[0].container_role == ""

    # -----------------------------------------------------------------------
    # Non-library components are skipped
    # -----------------------------------------------------------------------

    def test_non_library_components_skipped(self):
        sbom = _sbom(components=[
            {"type": "application", "name": "app", "version": "1.0", "purl": "pkg:pypi/app@1.0"},
            _comp(name="libfoo"),
        ])
        deps = self._extract(sbom)
        assert len(deps) == 1
        assert deps[0].name == "libfoo"

    # -----------------------------------------------------------------------
    # Image SBOM fixture (all fields together)
    # -----------------------------------------------------------------------

    def test_image_sbom_fixture(self):
        fixture = Path(__file__).parent / "fixtures" / "sbom" / "images" / "sbom-image-sample.json"
        sbom = json.loads(fixture.read_text())
        deps = self._extract(sbom)
        assert len(deps) == 3
        # All components come from an apk PURL
        for dep in deps:
            assert dep.package_type == "apk"
        # Container image name from metadata
        for dep in deps:
            assert dep.container_image == "sample-backend-image"


# ===========================================================================
# 4. Scanner parsers — VulnFinding.recommendation & acceptability_status
# ===========================================================================

class TestTrivyRecommendation:
    """trivy._parse() fills recommendation and acceptability_status."""

    def _parse(self, data: dict):
        import tempfile, json as _json
        from sbom_pipeline.scanner.trivy import _parse
        with tempfile.NamedTemporaryFile(suffix=".json", mode="w", delete=False) as f:
            _json.dump(data, f)
            p = Path(f.name)
        findings = _parse(p, "trivy")
        p.unlink(missing_ok=True)
        return findings

    def _trivy_result(self, **vuln_overrides) -> dict:
        vuln = {
            "VulnerabilityID": "CVE-2024-1111",
            "PkgName": "libssl",
            "InstalledVersion": "1.1.1k",
            "PkgRef": "pkg:apk/libssl@1.1.1k",
            "Severity": "HIGH",
            "Title": "OpenSSL heap overflow",
            "FixedVersion": "1.1.1n",
            "CVSS": {},
        }
        vuln.update(vuln_overrides)
        return {"Results": [{"Vulnerabilities": [vuln]}]}

    def test_recommendation_from_primary_url(self):
        data = self._trivy_result(PrimaryURL="https://nvd.nist.gov/vuln/detail/CVE-2024-1111")
        findings = self._parse(data)
        assert findings[0].recommendation == "https://nvd.nist.gov/vuln/detail/CVE-2024-1111"

    def test_recommendation_fallback_to_fixed_version(self):
        data = self._trivy_result()  # no PrimaryURL
        findings = self._parse(data)
        assert findings[0].recommendation == "Обновить до версии 1.1.1n"

    def test_recommendation_empty_when_no_fixed_no_url(self):
        data = self._trivy_result(FixedVersion="")
        findings = self._parse(data)
        assert findings[0].recommendation == ""

    def test_acceptability_status_populated(self):
        data = self._trivy_result(Status="fixed")
        findings = self._parse(data)
        assert findings[0].acceptability_status == "fixed"

    def test_acceptability_status_empty_when_absent(self):
        data = self._trivy_result()
        findings = self._parse(data)
        assert findings[0].acceptability_status == ""

    def test_acceptability_status_will_not_fix(self):
        data = self._trivy_result(Status="will_not_fix")
        findings = self._parse(data)
        assert findings[0].acceptability_status == "will_not_fix"


class TestClairRecommendation:
    """clair._parse() fills recommendation from Links."""

    def _parse(self, data: dict):
        import tempfile, json as _json
        from sbom_pipeline.scanner.clair import _parse
        with tempfile.NamedTemporaryFile(suffix=".json", mode="w", delete=False) as f:
            _json.dump(data, f)
            p = Path(f.name)
        findings = _parse(p)
        p.unlink(missing_ok=True)
        return findings

    def _clair_data(self, links: list | None = None) -> dict:
        return {
            "vulnerabilities": {
                "CVE-2024-2222": {
                    "Package": {"Name": "curl", "Version": "7.64.0"},
                    "NormalizedSeverity": "High",
                    "Description": "URL confusion in curl",
                    "FixedInVersion": "8.0.0",
                    "Links": links or [],
                }
            }
        }

    def test_recommendation_first_link(self):
        findings = self._parse(self._clair_data(links=["https://cve.mitre.org/cgi-bin/cvename.cgi?name=CVE-2024-2222", "https://other.link"]))
        assert findings[0].recommendation == "https://cve.mitre.org/cgi-bin/cvename.cgi?name=CVE-2024-2222"

    def test_recommendation_empty_when_no_links(self):
        findings = self._parse(self._clair_data(links=[]))
        assert findings[0].recommendation == ""

    def test_recommendation_empty_when_links_absent(self):
        data = {
            "vulnerabilities": {
                "CVE-2024-2222": {
                    "Package": {"Name": "curl", "Version": "7.64.0"},
                    "NormalizedSeverity": "High",
                    "Description": "desc",
                    "FixedInVersion": "",
                }
            }
        }
        findings = self._parse(data)
        assert findings[0].recommendation == ""


class TestDepcheckRecommendation:
    """depcheck._parse() fills recommendation from notes / references."""

    def _parse(self, data: dict):
        import tempfile, json as _json
        from sbom_pipeline.scanner.depcheck import _parse
        with tempfile.NamedTemporaryFile(suffix=".json", mode="w", delete=False) as f:
            _json.dump(data, f)
            p = Path(f.name)
        findings = _parse(p)
        p.unlink(missing_ok=True)
        return findings

    def _depcheck_data(self, notes: str = "", refs: list | None = None) -> dict:
        vuln: Dict[str, Any] = {
            "name": "CVE-2024-3333",
            "severity": "MEDIUM",
            "description": "Outdated library",
            "cvssv3": {"baseScore": 5.5},
        }
        if notes:
            vuln["notes"] = notes
        if refs is not None:
            vuln["references"] = refs
        return {
            "dependencies": [{
                "fileName": "log4j-1.2.jar",
                "packages": [{"id": "pkg:maven/log4j/log4j@1.2.17"}],
                "vulnerabilities": [vuln],
            }]
        }

    def test_recommendation_from_notes(self):
        findings = self._parse(self._depcheck_data(notes="Upgrade to log4j 2.x"))
        assert findings[0].recommendation == "Upgrade to log4j 2.x"

    def test_recommendation_from_references_url(self):
        refs = [{"url": "https://nvd.nist.gov/vuln/detail/CVE-2024-3333", "name": "NVD"}]
        findings = self._parse(self._depcheck_data(refs=refs))
        assert findings[0].recommendation == "https://nvd.nist.gov/vuln/detail/CVE-2024-3333"

    def test_notes_takes_precedence_over_references(self):
        refs = [{"url": "https://nvd.nist.gov/vuln/detail/CVE-2024-3333"}]
        findings = self._parse(self._depcheck_data(notes="Upgrade now", refs=refs))
        assert findings[0].recommendation == "Upgrade now"

    def test_recommendation_empty_when_no_notes_no_refs(self):
        findings = self._parse(self._depcheck_data())
        assert findings[0].recommendation == ""

    def test_recommendation_empty_ref_without_url(self):
        refs = [{"name": "NVD"}]  # no url key
        findings = self._parse(self._depcheck_data(refs=refs))
        assert findings[0].recommendation == ""


# ===========================================================================
# 5. VulnFinding dataclass — new fields present and default correctly
# ===========================================================================

class TestVulnFindingNewFields:
    def test_recommendation_default_empty(self):
        v = _vuln()
        assert v.recommendation == ""

    def test_acceptability_status_default_empty(self):
        v = _vuln()
        assert v.acceptability_status == ""

    def test_fields_set_correctly(self):
        v = _vuln(recommendation="Upgrade to 2.0", acceptability_status="acceptable")
        assert v.recommendation == "Upgrade to 2.0"
        assert v.acceptability_status == "acceptable"


# ===========================================================================
# 6. Exporter — _comp_rows() and _vuln_rows()
# ===========================================================================

class _FakeDep:
    """Minimal dependency object with all new attributes."""
    def __init__(self, **kwargs):
        self.name = kwargs.get("name", "libfoo")
        self.version = kwargs.get("version", "1.0.0")
        self.purl = kwargs.get("purl", "pkg:pypi/libfoo@1.0.0")
        self.package_type = kwargs.get("package_type", "pypi")
        self.srcLangs = kwargs.get("srcLangs", ["Python"])
        self.attack_surface = kwargs.get("attack_surface", "")
        self.security_function = kwargs.get("security_function", "")
        self.container_image = kwargs.get("container_image", "")
        self.container_role = kwargs.get("container_role", "")
        self.source = kwargs.get("source", "https://pypi.org/project/libfoo/1.0.0/")
        self.depType = kwargs.get("depType", [])


class TestExporterCompRows:
    """Exporter._comp_rows() maps all new component columns."""

    def _rows(self, **dep_kwargs) -> list:
        dep = _FakeDep(**dep_kwargs)
        exporter = Exporter([dep], vulns=[])
        return exporter._comp_rows()

    def test_package_type_column(self):
        rows = self._rows(package_type="maven")
        assert rows[0]["Тип пакета / тип компонента"] == "maven"

    def test_purl_column(self):
        rows = self._rows(purl="pkg:npm/lodash@4.17.21")
        assert rows[0]["PURL / технический идентификатор компонента"] == "pkg:npm/lodash@4.17.21"

    def test_attack_surface_column(self):
        rows = self._rows(attack_surface="yes")
        assert rows[0]["Признак принадлежности к поверхности атаки"] == "yes"

    def test_attack_surface_empty_by_default(self):
        rows = self._rows()
        assert rows[0]["Признак принадлежности к поверхности атаки"] == ""

    def test_security_function_column(self):
        rows = self._rows(security_function="tls")
        assert rows[0]["Признак выполнения функций безопасности"] == "tls"

    def test_security_function_empty_by_default(self):
        rows = self._rows()
        assert rows[0]["Признак выполнения функций безопасности"] == ""

    def test_container_image_column(self):
        rows = self._rows(container_image="my-image:1.0")
        assert rows[0]["Принадлежность к контейнерному образу"] == "my-image:1.0"

    def test_container_image_empty_for_git_sbom(self):
        rows = self._rows(container_image="")
        assert rows[0]["Принадлежность к контейнерному образу"] == ""

    def test_container_role_column(self):
        rows = self._rows(container_role="os-packages")
        assert rows[0]["Роль компонента в составе контейнерного образа"] == "os-packages"

    def test_container_role_empty_by_default(self):
        rows = self._rows()
        assert rows[0]["Роль компонента в составе контейнерного образа"] == ""

    def test_web_address_from_source(self):
        rows = self._rows(source="https://pypi.org/project/libfoo/")
        assert rows[0]["Адрес веб-ресурса"] == "https://pypi.org/project/libfoo/"

    def test_all_expected_columns_present(self):
        rows = self._rows()
        for col in _COMP_COLUMNS:
            assert col in rows[0], f"Missing column: {col}"

    def test_row_number_increments(self):
        dep1 = _FakeDep(name="a")
        dep2 = _FakeDep(name="b")
        exporter = Exporter([dep1, dep2], vulns=[])
        rows = exporter._comp_rows()
        assert rows[0]["№ п/п"] == 1
        assert rows[1]["№ п/п"] == 2

    def test_old_attack_surface_column_no_longer_present(self):
        """The merged column must NOT exist anymore."""
        rows = self._rows()
        assert "Принадлежность к поверхности атаки / функциям безопасности" not in rows[0]


class TestExporterVulnRows:
    """Exporter._vuln_rows() maps all new vulnerability columns."""

    def _rows(self, **vuln_kwargs) -> list:
        v = _vuln(**vuln_kwargs)
        exporter = Exporter([], vulns=[v])
        return exporter._vuln_rows()

    def test_recommendation_column_present(self):
        rows = self._rows(recommendation="Upgrade to 2.0")
        assert rows[0]["Рекомендация / компенсирующая мера"] == "Upgrade to 2.0"

    def test_recommendation_empty_by_default(self):
        rows = self._rows()
        assert rows[0]["Рекомендация / компенсирующая мера"] == ""

    def test_acceptability_status_column_present(self):
        rows = self._rows(acceptability_status="will_not_fix")
        assert rows[0]["Статус допустимости в рассматриваемой конфигурации"] == "will_not_fix"

    def test_acceptability_status_empty_by_default(self):
        rows = self._rows()
        assert rows[0]["Статус допустимости в рассматриваемой конфигурации"] == ""

    def test_all_expected_columns_present(self):
        rows = self._rows()
        for col in _VULN_COLUMNS:
            assert col in rows[0], f"Missing column: {col}"

    def test_existing_columns_still_present(self):
        rows = self._rows()
        assert rows[0]["CVE / ID"] == "CVE-2024-0001"
        assert rows[0]["Критичность"] == "HIGH"
        assert rows[0]["Исправлено в версии"] == "1.1.0"


# ===========================================================================
# 7. End-to-end: Excel export contains all new column headers
# ===========================================================================

class TestExcelExportHeaders:
    """The xlsx file produced by exportToExcel includes every new column."""

    def test_all_comp_columns_in_excel(self, tmp_path):
        import openpyxl

        dep = _FakeDep(
            package_type="pypi",
            purl="pkg:pypi/libfoo@1.0.0",
            attack_surface="yes",
            security_function="tls",
            container_image="my-image:1.0",
            container_role="app-layer",
        )
        out = str(tmp_path / "report.xlsx")
        Exporter([dep], vulns=[]).exportToExcel(out)

        wb = openpyxl.load_workbook(out)
        headers = [ws.cell(1, c).value for ws in wb.worksheets for c in range(1, ws.max_column + 1) if ws.cell(1, c).value]

        for col in _COMP_COLUMNS:
            assert col in headers, f"Missing in Excel: {col}"

    def test_all_vuln_columns_in_excel(self, tmp_path):
        import openpyxl

        v = _vuln(recommendation="Upgrade", acceptability_status="fixed")
        out = str(tmp_path / "report.xlsx")
        Exporter([_FakeDep()], vulns=[v]).exportToExcel(out)

        wb = openpyxl.load_workbook(out)
        vuln_ws = wb["Уязвимости"]
        headers = [vuln_ws.cell(1, c).value for c in range(1, vuln_ws.max_column + 1)]

        for col in _VULN_COLUMNS:
            assert col in headers, f"Missing in Excel vuln sheet: {col}"

    def test_vuln_data_row_contains_new_fields(self, tmp_path):
        import openpyxl

        v = _vuln(recommendation="Обновить до 2.0", acceptability_status="end_of_life")
        out = str(tmp_path / "report.xlsx")
        Exporter([_FakeDep()], vulns=[v]).exportToExcel(out)

        wb = openpyxl.load_workbook(out)
        ws = wb["Уязвимости"]
        # Row 1 = header, row 2 = first data row
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        rec_col = headers["Рекомендация / компенсирующая мера"]
        status_col = headers["Статус допустимости в рассматриваемой конфигурации"]
        assert ws.cell(2, rec_col).value == "Обновить до 2.0"
        assert ws.cell(2, status_col).value == "end_of_life"
